import openpyxl
from openpyxl.styles import PatternFill

class XlHandler:

    def __init__(self, path):
        self.wbObj = openpyxl.load_workbook(path)
        self.buySheet= self.wbObj.get_sheet_by_name("BUY")
        self.sellSheet = self.wbObj.get_sheet_by_name("SELL")
        self.redFill = PatternFill(bgColor="FF0000", fgColor="FF0000", fill_type="solid")
        self.greenFill = PatternFill(bgColor="00FF00",fgColor="00FF00", fill_type="solid")
        self.path = path

    def readBuyList(self):
        buyList = []
        for row in range(2,self.buySheet.max_row+1):
            buyList.append({
                "cardName": self.buySheet.cell(row,1).value,
                "set": self.buySheet.cell(row,2).value,
                "target": self.buySheet.cell(row,3).value,
                "trend": 0,
                "OK": False
            })
        return buyList

    def readSellList(self):
        sellList = []
        for row in range(2,self.sellSheet.max_row+1):
            sellList.append({
                "cardName": self.sellSheet.cell(row,1).value,
                "set": self.sellSheet.cell(row,2).value,
                "target": self.sellSheet.cell(row,3).value,
                "trend": 0,
                "OK": False
            })
        return sellList

    def writeBuyList(self, buyList):
        rowIndex = 2
        for element in buyList:
            self.buySheet.cell(rowIndex, 4).value = element["trend"]
            self.buySheet.cell(rowIndex, 5).value = element["OK"]
            if element["OK"]:
                self.buySheet.cell(rowIndex, 5).fill = self.greenFill
            else:
                self.buySheet.cell(rowIndex, 5).fill = self.redFill
            rowIndex += 1

    def writeSellList(self, sellList):
        rowIndex = 2
        for element in sellList:
            self.sellSheet.cell(rowIndex, 4).value = element["trend"]
            self.sellSheet.cell(rowIndex, 5).value = element["OK"]
            if element["OK"]:
                self.sellSheet.cell(rowIndex, 5).fill = self.greenFill
            else:
                self.sellSheet.cell(rowIndex, 5).fill = self.redFill
            rowIndex += 1

    def saveChanges(self):
        self.wbObj.save(self.path)